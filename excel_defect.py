#!/usr/bin/env python3
# coding: utf-8
from pathlib import Path
import time
from bigml.api import BigML
import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
import pickle

from tqdm import tqdm
from bigml_mining import get_or_create_association, get_or_create_dataset

from data_utils import read_dataset

def comparison(word):
	regexp =re.finditer(r'\w+=[0-9]+',word)
	regexp_minus_f = re.finditer(r'-[0-9]+.[0-9]+',word)
	regexp_minus = re.finditer(r'-[0-9]+(^.)',word)

	for m in regexp:
		matched_word=m.group()
		new_word="==".join(matched_word.split("="))
		word=word.replace(matched_word, new_word)
	return word


def import_excel(fileName, sheetName, title, dataframe):
    # Create a new workbook and add a worksheet
    book = Workbook()
    book.create_sheet(sheetName)
    
    # Acquire the sheet by its name
    sheet = book[sheetName]
    
    # Writing the title
    sheet.cell(row=1, column=1).value = title
    
    # Writing the dataframe to sheet
    rows = dataframe_to_rows(dataframe, index=False, header=True)
    for r_idx, row in enumerate(rows, 2):  # start from row 2 to leave room for the title
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Save the workbook
    book.save(fileName + '.xlsx')


def print_df(df_name,row_pos,col_pos,sheet_name):
	rows = dataframe_to_rows(df_name)
	
	for r_idx, row in enumerate(rows, row_pos):
		for c_idx, value in enumerate(row, col_pos):
			 sheet_name.cell(row=r_idx, column=c_idx, value=value)
	

def checkSheet(file_handler,sheet_name):
	if sheet_name in file_handler.sheetnames:
		sheet_name = file_handler[sheet_name]
	else:
		file_handler.create_sheet(sheet_name)#don't give extension
		sheet_name = file_handler[sheet_name]
	return sheet_name


def checkFile(path,filename):
	my_file = Path(path+filename)
	if my_file.is_file():
		book = load_workbook(filename)
	else:
		book = Workbook()
	return book


def printExcel(path,file, case_df,ff_df):
	book=checkFile(path,str(file)+'.xlsx')

	ff_results =checkSheet(book,'ff_results')
	print_df(case_df,1,1,ff_results)
	print_df(ff_df,2,1,ff_results)

	book.save(path+'/'+str(file)+'.xlsx')

def generate_excels(project):
	username = os.environ['BIGML_USERNAME']
	api_key = os.environ['BIGML_API_KEY']
	api = BigML(username, api_key)

	projects = read_dataset()
	models_path = Path("./models")

	train, test, val, inverse = projects[project]
	model_path = models_path / f"{project}.pkl"
	with open(model_path, "rb") as f:
		blackbox = pickle.load(f)

	generated_path = "./output/generated/" + project
	rules_path = Path("./rules") / project
	rules_path.mkdir(parents=True, exist_ok=True)
	output_path = Path("./output/SQAPlanner") / project
	output_path.mkdir(parents=True, exist_ok=True)


	for csv in tqdm(Path(generated_path).glob("*.csv"), desc="csv", leave=False, total=len(list(Path(generated_path).glob("*.csv")))):
		case_data = test.loc[int(csv.stem), :]
		case_data['target'] = case_data['target'].astype(int)
		case_data['target'] = case_data['target'].astype(str)
		x_test = case_data.drop("target")
		real_target = blackbox.predict(x_test.values.reshape(1, -1))
		if case_data['target'] == 0 or real_target == 0:
			continue

		if Path.exists(output_path / f"{csv.stem}.xlsx"):
			continue
		dataset_id = get_or_create_dataset(api, str(csv), project)
		options = {
			'name': csv.stem,
			'tags': [project],
			'search_strategy': 'confidence',
			'max_k': 30,
			'rhs_predicate': [{"field": "target", "operator": "=", "value": "False"}]
		}
		file = rules_path / csv
		get_or_create_association(api, dataset_id, options, str(file))

		ff_df= pd.DataFrame([])

		real_target= 'target=='+str(real_target[0])+str('.000')
		# Reading the file content to create a DataFrame
		rules_df = pd.read_csv(file, encoding='utf-8')#rule file

		for  index, row in rules_df.iterrows():
			rule = rules_df.iloc[index, 1]
			rule = comparison(rule)
			
			class_val = rules_df.iloc[index, 2]
			class_val = comparison(class_val)

			if real_target==class_val:
				print("correctly predicted")
			
			#Practices  to  follow  to  decrease  the  risk  of  having defects
			elif x_test.eval(rule).all()==False and real_target!=class_val:
				ff_df = pd.concat([ff_df,row.to_frame().T])

		printExcel(output_path, f"{csv.stem}.xlsx", x_test, ff_df)
				
	# Setting the file name (without extension) as the index name
def main():
	generate_excels('camel@0')


if __name__ == '__main__':
	main()
	# count = 0
	# while True:
	# 	try:
	# 		print(f"Running... ({count})")
	# 		main()
	# 		break
	# 	except Exception as e:
	# 		print(e)
	# 		print("Error occurred. Restarting...")
	# 		time.sleep(10)
	# 		count += 1


